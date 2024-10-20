import numpy as np
import csv
from openpyxl import load_workbook, Workbook
import json
import os
import pandas as pd


def csvRead(csv_file):
    # read csv file    
    f = open(csv_file, 'r')
    csv_reader = csv.reader(f)
    return_data = []
    for row in csv_reader:
        return_data.append(row)
    f.close()
    return np.array(return_data)


def csvSave(csv_file, save_data):
    # save csv file
    f = open(csv_file, 'w', newline='')
    csv_writer = csv.writer(f)
    for i in range(len(save_data)):
        csv_writer.writerow(save_data[i])
    f.close()


def xlsxSave0(xlsx_file, save_data, col_name, sheet_name='Sheet1'):
    data = dict()
    for col in range(len(col_name)):
        data[col_name[col]] = save_data[:, col]
    writeData = pd.DataFrame(data)
    if not os.path.exists(xlsx_file):
        writer = pd.ExcelWriter(xlsx_file, engine='openpyxl')
        writeData.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        writer = pd.ExcelWriter(xlsx_file, engine='openpyxl')
        book = load_workbook(xlsx_file)
        writer.book = book
        writeData.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()
    writer.close()


def xlsxSave(xlsx_file, save_data, col_name, sheet_name='Sheet1'):
    if not os.path.exists(xlsx_file):
        wb = Workbook()
        sh = wb.create_sheet(sheet_name)
    else:
        wb = load_workbook(xlsx_file)
        sh = wb.create_sheet(sheet_name)
    
    for col in range(len(col_name)):
        sh.cell(1, 1 + col).value = col_name[col]
    
    for row in range(len(save_data)):
        for col in range(len(col_name)):
            sh.cell(2 + row, 1 + col).value = save_data[row][col]
    wb.save(xlsx_file)


def jsonRead(json_file):
    # read json file
    f = open(json_file, 'r', encoding='UTF-8')
    dictionary = json.load(f)
    f.close()
    return dictionary


def jsonSave(save_file, dict_data):
    # save json file
    f = open(save_file, 'w')
    json.dump(dict_data, f, indent=4)
    f.close()


def txtRead(txt_file):
    # read txt file
    f = open(txt_file, 'r')
    txt_data = []
    data = f.readline()
    while data != '' and data != '\n':
        txt_data.append(data)
        data = f.readline()
    f.close()
    return txt_data


def txtSave(txt_file, save_data):
    # save txt file
    f = open(txt_file, 'w')
    for data in save_data:
        data += '\n'
        f.writelines(data)
    f.close()


def txtSaveAdd(txt_file, save_data):
    # save txt file - additional mode
    f = open(txt_file, 'a')
    for data in save_data:
        data += '\n'
        f.writelines(data)
    f.close()


def removeFolds(fold):
    files = os.listdir(fold)
    for filename in files:
        file = '{}/{}'.format(fold, filename)
        os.remove(file)
    os.rmdir(fold)
